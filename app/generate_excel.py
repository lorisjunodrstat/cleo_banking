import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel(stats, annee):
    """
    Génère un fichier Excel en mémoire à partir des données du compte de résultat.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Compte de résultat {annee}"
    
    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Titre du document
    ws['A1'] = f"COMPTE DE RÉSULTAT - ANNÉE {annee}"
    ws['A1'].font = font_title
    
    # En-têtes de tableau
    headers = ["Poste / Rubrique", "Montant (€)"]
    ws.append([]) # Ligne vide A2
    ws.append(headers) # Ligne A3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Insertion des données (adaptez les clés selon la structure de `stats`)
    # Exemple si `stats` contient des dictionnaires ou objets :
    produits = stats.get('produits', 0) if isinstance(stats, dict) else getattr(stats, 'produits', 0)
    charges = stats.get('charges', 0) if isinstance(stats, dict) else getattr(stats, 'charges', 0)
    resultat = produits - charges

    ws.append(["Total des Produits", produits])
    ws.append(["Total des Charges", charges])
    
    # Ligne Total / Résultat
    row_res = ws.max_row + 1
    ws.append(["Résultat Net", resultat])
    
    # Formatting des cellules de données
    for r in range(4, ws.max_row + 1):
        ws.cell(row=r, column=1).border = border_thin
        cell_val = ws.cell(row=r, column=2)
        cell_val.border = border_thin
        cell_val.number_format = '#,##0.00 €'
        
    # Style de la ligne de résultat final
    ws.cell(row=row_res, column=1).font = font_bold
    ws.cell(row=row_res, column=1).fill = fill_total
    ws.cell(row=row_res, column=2).font = font_bold
    ws.cell(row=row_res, column=2).fill = fill_total

    # Ajustement largeur des colonnes
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    # Sauvegarde dans un buffer mémoire
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()